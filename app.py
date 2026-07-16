import os
import datetime
from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = os.urandom(24)

RANKING_FILE = "ranking.xlsx"
ADMIN_PASSWORD = "jose2026"
INVITE_OPTIONS = {
    "homem": {
        "role": "padrinho",
        "label": "Você aceita ser meu padrinho?"
    },
    "mulher": {
        "role": "madrinha",
        "label": "Você aceita ser minha madrinha?"
    }
}

QUESTIONS = [
    {
        "question": "Com quantos centímetros eu nasci?",
        "choices": ["46 cm", "48 cm", "50 cm"],
        "answer": "48 cm"
    },
    {
        "question": "Quanto eu pesei quando nasci?",
        "choices": ["2 kg e 980 g", "3 kg e 68 g", "3 kg e 250 g"],
        "answer": "3 kg e 68 g"
    },
    {
        "question": "Que horas eu nasci?",
        "choices": ["14h", "16h", "18h"],
        "answer": "16h"
    },
    {
        "question": "Quantas furadinhas levei no teste do pezinho?",
        "choices": ["3", "4", "5"],
        "answer": "4"
    },
    {
        "question": "Quem me deu meu primeiro presente?",
        "choices": ["Tia Dete", "Prima Laura", "Vovó Maninha"],
        "answer": "Prima Laura"
    },
    {
        "question": "Quem me deu meu primeiro banho?",
        "choices": ["Tia Dete", "Vovó Maninha", "Vovó Patrícia"],
        "answer": "Tia Dete"
    },
    {
        "question": "Qual é meu nome completo?",
        "choices": ["José Leonardo Sousa Lima", "José Leonardo Lima de Sousa", "José L. de Sousa"],
        "answer": "José Leonardo Lima de Sousa"
    },
    {
        "question": "Em que dia descobriram que eu estava a caminho?",
        "choices": ["30/10/2024", "02/11/2024", "05/11/2024"],
        "answer": "02/11/2024"
    },
    {
        "question": "Quantos mL tomei na minha primeira mamadeira?",
        "choices": ["50 mL", "60 mL", "70 mL"],
        "answer": "60 mL"
    },
    {
        "question": "Qual foi meu primeiro remédio?",
        "choices": ["Simeticona", "Multi B Gotas", "Vitamina D"],
        "answer": "Multi B Gotas"
    }
]


def ensure_ranking_file():
    if not os.path.exists(RANKING_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ranking"
        sheet.append(["Nome", "Pontuação", "Acertos", "Erros", "Tempo", "Data", "Hora", "Convite Aceito"])
        workbook.save(RANKING_FILE)


def format_seconds_to_time(total_seconds):
    minutes = total_seconds // 60
    seconds = total_seconds % 60
    return f"{minutes:02d}:{seconds:02d}"


def parse_time_string(time_string):
    try:
        parts = time_string.split(":")
        if len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
        return int(time_string)
    except Exception:
        return 0


def normalize_score(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def write_ranking_entry(name, score, correct, wrong, time_str, accepted="Não"):
    ensure_ranking_file()
    workbook = load_workbook(RANKING_FILE, data_only=True)
    sheet = workbook.active
    timestamp = datetime.datetime.now()
    sheet.append([
        name,
        score,
        correct,
        wrong,
        time_str,
        timestamp.strftime("%d/%m/%Y"),
        timestamp.strftime("%H:%M:%S"),
        accepted
    ])
    row_index = sheet.max_row
    workbook.save(RANKING_FILE)
    return row_index


def update_ranking_acceptance(row_index):
    if not os.path.exists(RANKING_FILE):
        return False
    workbook = load_workbook(RANKING_FILE, data_only=True)
    sheet = workbook.active
    if row_index <= 1 or row_index > sheet.max_row:
        return False
    sheet.cell(row=row_index, column=8).value = "Sim"
    workbook.save(RANKING_FILE)
    return True


def get_ranking_entries():
    ensure_ranking_file()
    workbook = load_workbook(RANKING_FILE, data_only=True)
    sheet = workbook.active
    entries = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        name, score, correct, wrong, time_str, date_str, hour_str, accepted = row[:8]
        entries.append({
            "row_index": row_index,
            "name": name or "",
            "score": normalize_score(score),
            "correct": correct or 0,
            "wrong": wrong or 0,
            "time": time_str or "00:00",
            "date": date_str or "",
            "hour": hour_str or "",
            "accepted": accepted or "Não",
            "sort_time": parse_time_string(time_str or "00:00")
        })

    entries.sort(key=lambda item: (-item["score"], item["sort_time"], item["name"].lower()))
    ranked_entries = []
    for position, entry in enumerate(entries, start=1):
        ranked_entry = dict(entry)
        ranked_entry["position"] = position
        ranked_entries.append(ranked_entry)
    return ranked_entries


def get_top_ranking(limit=4):
    return get_ranking_entries()[:limit]


def get_participant_by_id(participant_id):
    for participant in get_ranking_entries():
        if participant["row_index"] == participant_id:
            return participant
    return None


def update_ranking_entry(participant_id, name, score):
    if not os.path.exists(RANKING_FILE):
        return False
    workbook = load_workbook(RANKING_FILE, data_only=True)
    sheet = workbook.active
    if participant_id <= 1 or participant_id > sheet.max_row:
        return False
    sheet.cell(row=participant_id, column=1).value = name
    sheet.cell(row=participant_id, column=2).value = score
    workbook.save(RANKING_FILE)
    return True


def delete_ranking_entry(participant_id):
    if not os.path.exists(RANKING_FILE):
        return False
    workbook = load_workbook(RANKING_FILE, data_only=True)
    sheet = workbook.active
    if participant_id <= 1 or participant_id > sheet.max_row:
        return False
    sheet.delete_rows(participant_id)
    workbook.save(RANKING_FILE)
    return True


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/audio")
def audio_player():
    return render_template("audio.html")


@app.route("/ranking")
def ranking():
    top_entries = get_top_ranking(limit=10)
    return render_template("ranking.html", top_entries=top_entries)


@app.route("/como-funciona")
def como_funciona():
    return render_template("como_funciona.html")


@app.route("/prosseguir")
def prosseguir():
    return redirect(url_for("nome"))


@app.route("/certificado")
def certificado():
    name = session.get("player_name") or "Seu nome"
    gender_key = session.get("player_gender", "homem")
    invite_info = INVITE_OPTIONS.get(gender_key, INVITE_OPTIONS["homem"])
    return render_template(
        "certificado.html",
        name=name,
        invite_role=invite_info["role"],
        today=datetime.datetime.now().strftime("%d/%m/%Y")
    )


@app.route("/certificado/<path:nome>")
def certificado_por_nome(nome):
    decoded_name = nome.replace("+", " ")
    return render_template(
        "certificado.html",
        name=decoded_name or "Seu nome",
        invite_role="padrinho",
        today=datetime.datetime.now().strftime("%d/%m/%Y")
    )


@app.route("/admin/ranking", methods=["GET", "POST"])
def admin_ranking():
    if request.method == "POST":
        password = request.form.get("password", "")
        if password == ADMIN_PASSWORD:
            session["admin_authenticated"] = True
            return redirect(url_for("admin_ranking"))
        return render_template("admin_ranking.html", participants=get_ranking_entries(), error="Senha incorreta.")

    if session.get("admin_authenticated"):
        participants = get_ranking_entries()
        return render_template("admin_ranking.html", participants=participants)

    return render_template("admin_login.html")


@app.route("/logout_admin")
def logout_admin():
    session.pop("admin_authenticated", None)
    return redirect(url_for("admin_ranking"))


@app.route("/editar_participante/<int:participant_id>", methods=["GET", "POST"])
def editar_participante(participant_id):
    if not session.get("admin_authenticated"):
        return redirect(url_for("admin_ranking"))
    participant = get_participant_by_id(participant_id)
    if not participant:
        return redirect(url_for("admin_ranking"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        score_value = request.form.get("score", "0").strip()

        if not name:
            return render_template(
                "editar_participante.html",
                participant=participant,
                error="Informe um nome para o participante."
            )

        try:
            score = int(score_value)
        except ValueError:
            return render_template(
                "editar_participante.html",
                participant=participant,
                error="A pontuação precisa ser um número inteiro."
            )

        update_ranking_entry(participant_id, name, score)
        return redirect(url_for("admin_ranking"))

    return render_template("editar_participante.html", participant=participant)


@app.route("/excluir_participante/<int:participant_id>", methods=["POST", "GET"])
def excluir_participante(participant_id):
    delete_ranking_entry(participant_id)
    return redirect(url_for("admin_ranking"))


@app.route("/nome", methods=["GET", "POST"])
def nome():
    if request.method == "POST":
        name = request.form.get("player_name", "").strip()
        gender = request.form.get("player_gender", "")
        if not name:
            return render_template(
                "nome.html",
                error="Por favor, informe como deseja aparecer no ranking.",
                player_name=name,
                player_gender=gender
            )
        if gender not in INVITE_OPTIONS:
            return render_template(
                "nome.html",
                error="Por favor, escolha se é homem ou mulher.",
                player_name=name,
                player_gender=gender
            )
        session.clear()
        session["player_name"] = name
        session["player_gender"] = gender
        return redirect(url_for("quiz"))
    return render_template("nome.html", player_name="", player_gender="")


@app.route("/quiz")
def quiz():
    if not session.get("player_name"):
        return redirect(url_for("nome"))
    return render_template("quiz.html", questions=QUESTIONS, player_name=session["player_name"])


@app.route("/save_result", methods=["POST"])
def save_result():
    payload = request.get_json()
    if not payload:
        return jsonify({"success": False, "message": "Dados não recebidos."}), 400

    name = session.get("player_name")
    if not name:
        return jsonify({"success": False, "message": "Nome não encontrado."}), 400

    score = int(payload.get("score", 0))
    correct = int(payload.get("correct", 0))
    wrong = int(payload.get("wrong", 0))
    time_spent = payload.get("time", "00:00")

    row_index = write_ranking_entry(name, score, correct, wrong, time_spent, accepted="Não")
    session["quiz_score"] = score
    session["quiz_correct"] = correct
    session["quiz_wrong"] = wrong
    session["quiz_time"] = time_spent
    session["ranking_row"] = row_index
    return jsonify({"success": True})


@app.route("/resultado")
def resultado():
    if not session.get("player_name") or session.get("quiz_score") is None:
        return redirect(url_for("nome"))

    top_entries = get_top_ranking(limit=4)
    return render_template(
        "resultado.html",
        name=session["player_name"],
        score=session["quiz_score"],
        correct=session["quiz_correct"],
        wrong=session["quiz_wrong"],
        time=session["quiz_time"],
        top_entries=top_entries
    )


@app.route("/convite", methods=["GET", "POST"])
def convite():
    if not session.get("player_name") or session.get("ranking_row") is None or not session.get("player_gender"):
        return redirect(url_for("nome"))

    gender_key = session.get("player_gender", "homem")
    invite_info = INVITE_OPTIONS.get(gender_key, INVITE_OPTIONS["homem"])

    if request.method == "POST":
        row_index = session.get("ranking_row")
        success = update_ranking_acceptance(row_index)
        return jsonify({"success": success})

    return render_template(
        "convite.html",
        invite_label=invite_info["label"],
        invite_role=invite_info["role"]
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
